"""Tests for the shared Staff salary layout (parse_staff_sheet)."""
import openpyxl
import pytest
from vaishali.payroll.ingest.staff_layout import parse_staff_sheet, COLS


def _make_synth_workbook(tmp_path, rows: list[dict]):
    """Build a minimal Staff-shaped workbook from a list of row dicts.
    Each dict supplies whatever COLS keys the test cares about; missing
    keys get None. Returns the file path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synth Sheet"
    # rows 1-2: title (anything non-None) so DATA_START=4 row index is right
    ws.cell(row=1, column=1, value="title")
    ws.cell(row=2, column=1, value="subtitle")
    # row 3: header — copy COL keys verbatim into the cells
    for key, col in COLS.items():
        ws.cell(row=3, column=col, value=key)
    # rows 4+: data
    for i, row_data in enumerate(rows):
        for key, val in row_data.items():
            ws.cell(row=4 + i, column=COLS[key], value=val)
    out = tmp_path / "synth.xlsx"
    wb.save(out)
    return str(out)


def test_employee_named_subhash_not_dropped(tmp_path):
    """A real employee whose name starts with 'Subhash' must NOT be skipped
    by the Total/Sub-Total prefix filter. This is a regression guard against
    a substring `"Sub" in name` bug that was caught and fixed earlier."""
    xlsx = _make_synth_workbook(tmp_path, [
        {"sr": 1, "emp_code": "X1", "name": "Subhash Patel",
         "days_worked": 31, "gross_pay_payable": 25000},
        {"sr": 2, "emp_code": "X2", "name": "Subodh Kumar",
         "days_worked": 22, "gross_pay_payable": 18000},
        {"sr": 3, "emp_code": "X3", "name": "Total",
         "days_worked": None, "gross_pay_payable": 43000},
        {"sr": 4, "emp_code": "X4", "name": "Subtotal",
         "days_worked": None, "gross_pay_payable": 43000},
    ])
    rows = parse_staff_sheet(xlsx, "Synth Sheet")
    names = {r["name"] for r in rows}
    assert "Subhash Patel" in names, "Subhash must not be filtered out"
    assert "Subodh Kumar" in names, "Subodh must not be filtered out"
    assert "Total" not in names, "literal 'Total' row must be dropped"
    assert "Subtotal" not in names, "'Subtotal' row must be dropped"
    assert len(rows) == 2


def test_uncached_formula_guard_raises(tmp_path):
    """When days_worked > 0 but gross_pay_payable is None, parse_staff_sheet
    must raise ValueError with a message about Excel re-saving."""
    xlsx = _make_synth_workbook(tmp_path, [
        {"sr": 1, "emp_code": "X1", "name": "Asha Test",
         "days_worked": 31, "gross_pay_payable": None},  # the trap
    ])
    with pytest.raises(ValueError, match="uncached"):
        parse_staff_sheet(xlsx, "Synth Sheet")
