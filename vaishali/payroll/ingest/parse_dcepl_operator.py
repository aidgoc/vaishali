"""Parser for file 3 — DCEPL Operator Salary Register Mar 2026.

Two parts:
  - parse_subgroups()  — reads side sheets `OP` and `VB OP` to classify each
    operator emp_code as Standard or VB
  - parse()            — reads the main `OP Salary Mar 2026` sheet, enriches
    each row with payroll_subgroup
"""
from __future__ import annotations
import openpyxl

MAIN_SHEET = "OP Salary Mar 2026"
OP_SHEET = "OP"
VB_OP_SHEET = "VB OP"

# Main-sheet columns (1-based)
COLS = {
    "sr": 1, "emp_code": 2, "uan": 3, "pf_no": 4, "esic_no": 5,
    "hrc_remark": 6, "name": 7, "gender": 8,
    "cross_check": 9, "salary_wef": 10,
    "days_present": 11, "salary_gross_target": 12,
    "hr_compliance": 13, "pf_applicable": 14, "esic_applicable": 15,
    "ex_gratia_applicable": 16, "member_sanchay": 17,
    "ot_applicable": 18, "group_medical": 19,
    "gross_payable": 20,
    "basic": 21, "da": 22, "da_arrears": 23, "hra": 24,
    "attendance_allow": 25, "food_allow": 26, "sp_allow": 27,
    "total_allowances": 28, "gross_payable_present_days": 29,
    "ot_rate": 30, "ot_hours": 31, "ot_amount": 32,
    "salary_with_ot": 33, "site_allow": 34, "site_allow_formula": 35,
    "total_gross": 36, "pf": 37, "esi": 38, "pt": 39,
    "tour_advance": 40, "sal_advance": 41, "group_medical_deduct": 42,
    "pf_wo_da_arrear": 43, "pf_diff": 44,
    "esic_wo_da_arrears": 45, "esic_diff": 46,
    "mlwf": 47, "arrears": 48, "net_salary": 49,
}

DATA_START = 4  # main sheet


def parse_subgroups(xlsx_path: str) -> dict[str, str]:
    """Return {emp_code: 'Standard' | 'VB'} for every operator listed on the
    OP and VB OP side sheets."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: dict[str, str] = {}
    for sheet_name, label in ((OP_SHEET, "Standard"), (VB_OP_SHEET, "VB")):
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found in {xlsx_path}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        # Header row 1, data from row 2 onwards
        for r in range(2, ws.max_row + 1):
            emp_code = ws.cell(row=r, column=2).value
            if not emp_code:
                continue
            out[str(emp_code).strip()] = label
    return out


def parse(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if MAIN_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet {MAIN_SHEET!r} not found in {xlsx_path}. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[MAIN_SHEET]
    subgroups = parse_subgroups(xlsx_path)

    rows = []
    for r in range(DATA_START, ws.max_row + 1):
        sr = ws.cell(row=r, column=COLS["sr"]).value
        emp_code = ws.cell(row=r, column=COLS["emp_code"]).value
        name = ws.cell(row=r, column=COLS["name"]).value
        if sr is None or emp_code is None or name is None:
            continue
        emp_code = str(emp_code).strip()
        name_str = str(name).strip()
        if not emp_code:
            continue
        if name_str.lower().startswith(("total", "sub total", "subtotal", "grand total")):
            continue
        row = {key: ws.cell(row=r, column=col).value for key, col in COLS.items()}
        row["emp_code"] = emp_code
        row["name"] = name_str
        row["payroll_subgroup"] = subgroups.get(emp_code, "Standard")
        rows.append(row)
    return rows
