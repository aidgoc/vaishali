"""Parser for file 4 — Overhead Salary Register Mar 2026.

4 overhead employees (mixed DCEPL/DSPL/NA company tagging) on a 5-component
percentage-based structure:
  Basic + DA = 40%, HRA = 18.4%, Attendance = 15.2%, Food = 16%, SP = 10.4%
  No PF, no ESIC. MLWF only deduction.
"""
from __future__ import annotations
import openpyxl

SHEET = "Overhead Salary Mar 2026"
HEADER_ROW = 2
DATA_START = 8

# 1-based column indices, verified against actual Excel
COLS = {
    "sr": 1, "emp_code": 2, "uan": 3, "pf_no": 4, "esic_no": 5,
    "dept": 6, "company": 7, "name": 8,
    "cross_check": 9, "wef": 10,
    "days_present": 11, "salary_gross_target": 12,
    "pf_applicable": 13, "esic_applicable": 14, "ex_gratia_applicable": 15,
    "member_sanchay": 16, "ot_applicable": 17, "hr_compliance_7th": 18,
    "group_medical": 19,
    "gross_as_per_paid_days": 20,
    "basic_salary": 21, "hra": 22, "attendance_allow": 23,
    "food_allow": 24, "sp_allow": 25,
    "total_allowances": 26, "gross_payable_present_days": 27,
    "ot_rate": 28, "ot_hours": 29, "ot_amount": 30,
    "salary_with_ot": 31, "site_allow": 32,
    "total_gross": 33,
    "pf": 34, "esi": 35, "pt": 36, "tds": 37,
    "sal_advance": 38, "mlwf": 39,
    "uniform_deposit": 40, "credit_uniform_iou": 41,
    "net_salary": 42,
}


def parse(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet {SHEET!r} not found in {xlsx_path}. "
            f"Available sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET]
    rows = []
    for r in range(DATA_START, ws.max_row + 1):
        sr = ws.cell(row=r, column=COLS["sr"]).value
        name = ws.cell(row=r, column=COLS["name"]).value
        if sr is None or not name:
            continue
        name_str = str(name).strip()
        if name_str.lower().startswith(("total", "sub total", "subtotal", "grand total")):
            continue
        row = {key: ws.cell(row=r, column=col).value for key, col in COLS.items()}
        row["name"] = name_str
        rows.append(row)
    return rows
