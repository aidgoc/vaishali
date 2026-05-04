"""Parser for file 1 — DCEPL Employee Salary Register & Attendance Mar 2026.

Returns one dict per employee row. Column indices below are 1-based to match
openpyxl. The Excel has two layout quirks:
  - Header is on row 3 (rows 1-2 are merged title cells)
  - Total rows interleave with data — we skip any row whose col 1 (Sr) is None
"""
from __future__ import annotations
import openpyxl

SHEET = "DCEPL Salary Sheet-Mar 2026"
HEADER_ROW = 3
DATA_START = 4

COLS = {
    "sr": 1, "emp_code": 2, "uan": 3, "pf_no": 4, "esic_no": 5,
    "company": 6, "dept": 7, "name": 8, "gender": 9,
    "aadhar": 10, "pan": 11, "doj": 12, "exit_date": 13,
    "month_days": 14, "days_worked": 15,
    "ot_days": 16, "ot_amount_target": 17,
    "gross_pay_target": 18, "basic_target": 19, "hra_target": 20,
    "medical_target": 21, "lta_target": 22, "conv_target": 23,
    "chedu_target": 24, "food_target": 25, "stat_bonus_target": 26,
    "special_target": 27,
    "ot_amount_payable": 28, "gross_pay_target_check": 29,
    "basic_payable": 30, "hra_payable": 31,
    "medical_payable": 32, "lta_payable": 33, "conv_payable": 34,
    "chedu_payable": 35, "food_payable": 36, "stat_bonus_payable": 37,
    "special_payable": 38, "other_amount": 39, "gross_pay_payable": 40,
    "pf": 41, "esic": 42, "pt": 43, "mlwf": 44, "tds": 45,
    "other_deductions": 46, "salary_arrears": 47,
    "total_deduction": 48, "net_salary": 49,
}


def parse(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET]

    rows = []
    for r in range(DATA_START, ws.max_row + 1):
        sr = ws.cell(row=r, column=COLS["sr"]).value
        name = ws.cell(row=r, column=COLS["name"]).value
        if sr is None or name is None:
            continue
        name_str = str(name).strip()
        if name_str.lower().startswith(("total", "sub total", "subtotal", "grand total")):
            continue
        row = {key: ws.cell(row=r, column=col).value for key, col in COLS.items()}
        row["name"] = str(row["name"]).strip()
        if not row["emp_code"]:
            continue
        rows.append(row)

    # Guard against uncached formulas — openpyxl returns None for un-recalculated cells.
    # If a worker with days_worked > 0 shows None on gross_pay_payable, the cache is stale.
    for row in rows:
        if row.get("days_worked") and row.get("gross_pay_payable") is None:
            raise ValueError(
                f"Row for {row.get('name')!r}: gross_pay_payable is None but days_worked={row['days_worked']}. "
                "Excel formulas appear uncached. Open the file in Microsoft Excel and re-save before parsing."
            )
    return rows
