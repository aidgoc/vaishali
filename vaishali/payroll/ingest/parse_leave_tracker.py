"""Parser for file 7 — Dynamic Emp Leave Data Apr 2025 - Mar 2026.

Each employee block on the sheet looks like:
  row N+0: index (col 1) + emp_code (col 2, often blank) + name (col 3) + col 4 = 'LEAVES'
           + month headers APR..MAR in cols 6..17
  row N+1: Pre Leave Bal (per-month opening)
  row N+2: Earned Leaves (1.5/month accrual)
  row N+3: Taken Leaves
  (further rows — Remaining Bal / LWP — are present but not consumed here)

We compute closing bal for Mar = Pre[MAR] + Earned[MAR] − Taken[MAR],
clamped to non-negative (leave balance can't go below zero in this system).

Many rows in the source file have a blank emp_code (col 2). We fall back to
name (col 3) as the dict key in that case — Task 20 resolves either form to
an Employee via ERPNext lookup.
"""
from __future__ import annotations
import openpyxl

SHEETS = ["HO_Leave Data_2025-2026", "Pirangut_Leave Data_2025-2026"]
MAR_COL = 17  # April starts at col 6, March = col 6+11 = 17


def parse(xlsx_path: str) -> dict[str, float]:
    """Return {emp_code_or_name: mar_closing_bal} across HO + Pirangut sheets."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: dict[str, float] = {}
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found in {xlsx_path}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            v_col4 = ws.cell(row=r, column=4).value
            if not (v_col4 and "LEAVES" in str(v_col4)):
                continue
            v_col2 = ws.cell(row=r, column=2).value
            v_col3 = ws.cell(row=r, column=3).value
            key = str(v_col2).strip() if v_col2 else (str(v_col3).strip() if v_col3 else None)
            if not key:
                continue
            pre = ws.cell(row=r + 1, column=MAR_COL).value or 0
            earned = ws.cell(row=r + 2, column=MAR_COL).value or 0
            taken = ws.cell(row=r + 3, column=MAR_COL).value or 0
            closing = max(0.0, float(pre) + float(earned) - float(taken))
            out[key] = closing
    return out
